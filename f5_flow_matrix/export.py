"""Export de la matrice de flux en CSV et XLSX."""

from __future__ import annotations

import csv
from pathlib import Path

from .flow_matrix import COLUMNS, DATAGROUP_COLUMNS

FLOW_TYPE_COLORS = {
    "Ingress-LB": "D9E8FB",
    "Backend-LB": "E2F0D9",
    "Firewall-Rule": "FCE4D6",
    "GTM-DNS": "F2E2F5",
    "VIP-Orphan": "FFF2CC",
}

NEEDS_REVIEW_COLOR = "FFF2CC"

MAIN_COLUMN_WIDTHS = {
    "Device": 18,
    "Hostname": 22,
    "FlowType": 14,
    "Source": 28,
    "SourcePort": 10,
    "Destination": 28,
    "DestinationPort": 14,
    "Protocol": 10,
    "Action": 20,
    "ObjectType": 16,
    "ObjectName": 40,
    "Detail": 60,
    "NeedsReview": 45,
}

DATAGROUP_COLUMN_WIDTHS = {
    "Device": 18,
    "DataGroupName": 30,
    "Type": 10,
    "RecordKey": 35,
    "RecordValue": 30,
    "ReferencedByIRules": 40,
}


def write_csv(rows: list[dict], path: str | Path, columns: list[str] = COLUMNS) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _style_sheet(ws, columns: list[str], rows: list[dict], widths: dict, flow_type_col: str | None = None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(columns)
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    flow_type_idx = columns.index(flow_type_col) + 1 if flow_type_col else None
    review_idx = columns.index("NeedsReview") + 1 if "NeedsReview" in columns else None
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
        if flow_type_idx:
            color = FLOW_TYPE_COLORS.get(row.get(flow_type_col))
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.cell(row=ws.max_row, column=flow_type_idx).fill = fill
        if review_idx and row.get("NeedsReview"):
            fill = PatternFill(start_color=NEEDS_REVIEW_COLOR, end_color=NEEDS_REVIEW_COLOR, fill_type="solid")
            ws.cell(row=ws.max_row, column=review_idx).fill = fill

    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)


def write_xlsx(
    rows: list[dict],
    path: str | Path,
    datagroup_rows: list[dict] | None = None,
) -> None:
    from openpyxl import Workbook

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Matrice de flux"
    _style_sheet(ws_main, COLUMNS, rows, MAIN_COLUMN_WIDTHS, flow_type_col="FlowType")

    review_rows = [r for r in rows if r.get("NeedsReview")]
    ws_review = wb.create_sheet("Points d'attention")
    _style_sheet(ws_review, COLUMNS, review_rows, MAIN_COLUMN_WIDTHS, flow_type_col="FlowType")

    if datagroup_rows:
        ws_dg = wb.create_sheet("DataGroups (ref. iRules)")
        _style_sheet(ws_dg, DATAGROUP_COLUMNS, datagroup_rows, DATAGROUP_COLUMN_WIDTHS)

    wb.save(path)
